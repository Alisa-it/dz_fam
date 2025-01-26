from openpyxl import load_workbook

def from_variant(variant):
    wb = load_workbook('fam.xlsx')
    sheet = wb["Дано"]
    lst = []
    row_index = variant + 1
    for cellObj in sheet[f'A{row_index}':f'O{row_index}']:
        for cell in cellObj:
            lst.append(cell.value)
    wb.save("fam.xlsx")
    max_var = sheet.max_row - 1
    var_nums = []
    for i in range(max_var):
        var_nums.append(i+1)
    return lst, var_nums
    